"""
Violation Reporter
- RRP'nin 499 TL+ altındaki satışları tespit eder
- Mağazanın kendi sayfasının ekran görüntüsünü alır
- Excel raporuna kaydeder
"""

import os, re, json
from datetime import datetime
from pathlib import Path

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EVIDENCE_DIR = os.path.join(BASE_DIR, "evidence")
REPORT_FILE = os.path.join(BASE_DIR, "violations_report.xlsx")
VIOLATION_THRESHOLD = 499  # TL

os.makedirs(EVIDENCE_DIR, exist_ok=True)


# ── Excel report ──────────────────────────────────────────────

HEADERS = [
    "Tarih & Saat",
    "Ürün Adı",
    "Mağaza",
    "RRP (TL)",
    "Satış Fiyatı (TL)",
    "Fark (TL)",
    "Fark (%)",
    "Not",
    "Ekran Görüntüsü",
    "Mağaza URL",
]

HEADER_COLOR = "1A1916"
HEADER_FONT  = "FFFFFF"
ROW_COLORS   = ["FFFFFF", "F5F4F1"]
VIOLATION_COLOR = "FFE8E6"  # kırmızımsı satır
BELOW_COLOR     = "E8F5EE"  # yeşilimsi (sadece bilgi)


def _get_or_create_workbook():
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    if os.path.exists(REPORT_FILE):
        wb = load_workbook(REPORT_FILE)
        ws = wb.active
        return wb, ws

    wb = Workbook()
    ws = wb.active
    ws.title = "İhlal Raporu"

    # Header row
    from openpyxl.styles import Font, PatternFill, Alignment
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(name="Arial", bold=True, color=HEADER_FONT, size=10)
        cell.fill = PatternFill("solid", start_color=HEADER_COLOR)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Column widths
    widths = [18, 28, 18, 12, 16, 12, 10, 30, 22, 40]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = w

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    wb.save(REPORT_FILE)
    return wb, ws


def _add_violation_row(violation: dict):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = load_workbook(REPORT_FILE)
    ws = wb.active
    next_row = ws.max_row + 1
    is_odd = (next_row % 2 == 0)
    bg = VIOLATION_COLOR  # ihlal satırları hep kırmızımsı

    values = [
        violation["timestamp"],
        violation["product_name"],
        violation["store"],
        violation["rrp"],
        violation["price"],
        violation["diff"],
        violation["diff_pct"],
        "",  # Not — kullanıcı doldurur
        violation["screenshot_filename"] or "—",
        violation["store_url"],
    ]

    for col, val in enumerate(values, 1):
        cell = ws.cell(row=next_row, column=col, value=val)
        cell.font = Font(name="Arial", size=9)
        cell.fill = PatternFill("solid", start_color=bg)
        cell.alignment = Alignment(vertical="center", wrap_text=(col in [2, 8, 9]))

        # Sayısal sütunları formatla
        if col in [4, 5, 6]:
            cell.number_format = '#,##0'
        if col == 7:
            cell.number_format = '0.0"%"'

    ws.row_dimensions[next_row].height = 22
    wb.save(REPORT_FILE)


# ── Screenshot ────────────────────────────────────────────────

def _take_screenshot(url: str, filename: str) -> bool:
    """Mağazanın kendi URL'sine gidip ekran görüntüsü al"""
    try:
        from playwright.sync_api import sync_playwright
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True, args=["--no-sandbox"])
            ctx = browser.new_context(
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/122.0.0.0 Safari/537.36",
                viewport={"width": 1280, "height": 900},
                locale="tr-TR",
            )
            ctx.add_init_script("Object.defineProperty(navigator,'webdriver',{get:()=>undefined})")
            page = ctx.new_page()
            page.goto(url, wait_until="domcontentloaded", timeout=25000)
            page.wait_for_timeout(2500)
            path = os.path.join(EVIDENCE_DIR, filename)
            page.screenshot(path=path, full_page=True)
            browser.close()
        return True
    except Exception as e:
        print(f"    [Screenshot] Error: {e}")
        return False


def _resolve_store_url(akakce_url: str) -> str:
    """
    Akakce redirect URL'sini (akakce.com/c/?...) gerçek mağaza URL'sine çevirir.
    Playwright ile takip eder.
    """
    if not akakce_url or "akakce.com/c/" not in akakce_url:
        return akakce_url
    try:
        from playwright.sync_api import sync_playwright
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True, args=["--no-sandbox"])
            page = browser.new_context().new_page()
            page.goto(akakce_url, wait_until="domcontentloaded", timeout=15000)
            page.wait_for_timeout(2000)
            final_url = page.url
            browser.close()
        return final_url
    except:
        return akakce_url


# ── Main entry point ──────────────────────────────────────────

def process_violations(product_name: str, rrp: float, channels: list):
    """
    Scraper'dan gelen channel listesini kontrol eder.
    RRP - price > VIOLATION_THRESHOLD olan her satıcı için kayıt oluşturur.
    """
    violations_found = []

    # Excel yoksa oluştur
    _get_or_create_workbook()

    for ch in channels:
        price = ch.get("price")
        if not price or not rrp:
            continue
        diff = rrp - price
        if diff <= VIOLATION_THRESHOLD:
            continue

        diff_pct = round((diff / rrp) * 100, 1)
        store = ch.get("store", "?")
        store_url = ch.get("url", "")
        now = datetime.now()
        ts = now.strftime("%d.%m.%Y %H:%M")
        safe_name = re.sub(r'[^\w]', '_', f"{now.strftime('%Y%m%d_%H%M')}_{store}_{product_name[:20]}")
        screenshot_filename = f"{safe_name}.png"

        print(f"  [VİOLATION] {store}: {price:,.0f} TL (RRP: {rrp:,.0f}, Fark: {diff:,.0f} TL)")

        # Mağaza URL'sini çöz
        print(f"    → URL çözülüyor...")
        real_url = _resolve_store_url(store_url)
        print(f"    → {real_url[:60]}...")

        # Ekran görüntüsü al
        print(f"    → Ekran görüntüsü alınıyor...")
        ok = _take_screenshot(real_url, screenshot_filename)
        if not ok:
            screenshot_filename = None

        violation = {
            "timestamp": ts,
            "product_name": product_name,
            "store": store,
            "rrp": rrp,
            "price": price,
            "diff": round(diff, 2),
            "diff_pct": diff_pct,
            "screenshot_filename": screenshot_filename,
            "store_url": real_url,
        }

        _add_violation_row(violation)
        violations_found.append(violation)
        print(f"    ✓ Kayıt edildi → violations_report.xlsx")

    return violations_found